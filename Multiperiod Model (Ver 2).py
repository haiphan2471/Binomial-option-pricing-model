########################################################################################
## FileName:       Binominal Option Pricing Models
## FileType:       Python IDLE
## Author:         Hai Phan
## Adviser:        Seonguk Kim
## Course:         Summer Research 2021
## Institute:      DePauw University, IN, US
## Created In:     June - July, 2021
##
## Description:    The program uses input data from an existing xlsx file to
##                 calculate the pay of price for European Option and American Option,
##                 build the binomial trees for Stock Option, European Option, and American
##                 Option, then visualize the data with tables in the same xlsx file.
########################################################################################


import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


wb = load_workbook('Binomial Option Pricing Model.xlsx')
ws = wb['Stock Option']
S = ws['O2'].value
u = ws['O3'].value
d = ws['O4'].value
K = ws['O5'].value
r = ws['O6'].value
T = ws['O7'].value
n = ws['O8'].value


R = 1 + r
p = (R - d) / (u - d) #probability
step = n / T
stock = np.zeros((n+1, n+1)) 
ECpayoff = np.zeros((n+1, n+1)) #EC: Euro Call
EPpayoff = np.zeros((n+1, n+1)) #EP: Euro Put
ACpayoff = np.zeros((n+1, n+1)) #AC: American Call
APpayoff = np.zeros((n+1, n+1)) #AP: American Put


#___clear previous xlsx version___


ws.freeze_panes = None
ws_active = wb['Stock Option']

ws['O10'].value = None
ws['O11'].value = None
ws['O12'].value = None
ws['O13'].value = None
ws['O14'].value = None

for row in ws['A1':'L100']:
    for cell in row:
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
wb.save('Binomial Option Pricing Model.xlsx')


#___calc stock___


stock[0][0] = S
for i in range (1, n+1):
    stock[0][i] = stock[0][i-1] * d
for j in range (1, n+1):
    for i in range (1, j+1):
        stock[i][j] = stock[i-1][j-1] * u


#___calc euro call payoff___

        
for i in range (0, n+1):
    ECpayoff[i][n] = max(0, stock[i][n] - K)
for j in range (n-1, -1, -1):
    for i in range (j, -1, -1):
        ECpayoff[i][j] = (p * ECpayoff[i+1][j+1] + (1 - p) * ECpayoff[i][j+1]) / R
finalECpayoff = ECpayoff[0][0]


#___calc euro put payoff___


for i in range (0, n+1):
    EPpayoff[i][n] = max(0, K - stock[i][n])
for j in range (n-1, -1, -1):
    for i in range (j, -1, -1):
        EPpayoff[i][j] = (p * EPpayoff[i+1][j+1] + (1 - p) * EPpayoff[i][j+1]) / R
finalEPpayoff = EPpayoff[0][0]


#___calc american call payoff___


holding_value = 0.00
excercise_value = 0.00
for i in range (0, n+1):
    ACpayoff[i][n] = ECpayoff[i][n]
for j in range(n-1, -1, -1):
    for i in range(j, -1, -1):
        holding_value = (p * ACpayoff[i+1][j+1] + (1 - p) * ACpayoff[i][j+1]) / R
        excercise_value = max(0, stock[i][n] - K)
        ACpayoff[i][n] = max(holding_value, excercise_value)
finalACpayoff = ACpayoff[0][0]


#___calc american put payoff___


for i in range (0, n+1):
    APpayoff[i][n] = EPpayoff[i][n]
for j in range(n-1, -1, -1):
    for i in range(j, -1, -1):
        holding_value = (p * APpayoff[i+1][j+1] + (1 - p) * APpayoff[i][j+1]) / R
        excercise_value = max(0, K - stock[i][n])
        APpayoff[i][n] = max(holding_value, excercise_value)
finalAPpayoff = APpayoff[0][0]


#___print output value___
        

ws['O10'] = p
ws['O10'].font = Font(color='FF0000', bold=True)
ws['O11'] = finalECpayoff
ws['O11'].font = Font(color='FF0000', bold=True)
ws['O12'] = finalEPpayoff
ws['O12'].font = Font(color='FF0000', bold=True)
ws['O13'] = finalACpayoff
ws['O13'].font = Font(color='FF0000', bold=True)
ws['O14'] = finalAPpayoff
ws['O14'].font = Font(color='FF0000', bold=True)


#___print stock tree___(Orange)

align = Alignment(horizontal='center', vertical='center')

ws.cell(1,1, 'Stock Tree')
ws.cell(1,1).font = Font(bold=True)
for i in range(1, n+3): #title line
    ws.cell(1, i).fill = PatternFill(start_color='ff7f50', end_color='ff7f50', fill_type='solid')
ws.cell(n+3, 1).fill = PatternFill(start_color='ffb79d', end_color='ffb79d', fill_type='solid')
for j in range(n+1):
    for i in range(j+1):
        if j == 1:
            for k in range(n+1):#vertical line
                ws.cell(n+1-k+1,j,k)
                ws.cell(n+1-k+1,j).fill = PatternFill(start_color='ffb79d', end_color='ffb79d', fill_type='solid')
                ws.cell(n+1-k+1,j).alignment = align
                
        ws.cell(n+1-i+1,j+2,round(stock[i][j],4))
        ws.cell(n+1-i+1,j+2).alignment = align
        
        ws.cell(n+3,j+2,j)#horizonal line
        ws.cell(n+3,j+2).fill = PatternFill(start_color='ffb79d', end_color='ffb79d', fill_type='solid')
        ws.cell(n+3,j+2).alignment = align


#___print Euro Call tree___(Green)

        
ws.cell(n+5,1, 'Euro Call Payoff Tree')
ws.cell(n+5, 1).font = Font(bold=True)
for i in range(1, n+3):
    ws.cell(n+5, i).fill = PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid')
ws.cell(2*n+7, 1).fill = PatternFill(start_color='b8dc70', end_color='b8dc70', fill_type='solid')
for j in range(n+1):
    for i in range(j+1):
        if j == 1:
            for k in range(n+1):
                ws.cell(2*(n)+5-k+1,j,k)
                ws.cell(2*(n)+5-k+1,j).fill = PatternFill(start_color='b8dc70', end_color='b8dc70', fill_type='solid')
                ws.cell(2*(n)+5-k+1,j)
                
        ws.cell(2*(n)+5-i+1,j+2,round(ECpayoff[i][j],4))
        ws.cell(2*(n)+5-i+1,j+2).alignment = align
        
        ws.cell(2*(n)+7,j+2,j)
        ws.cell(2*(n)+7,j+2).fill = PatternFill(start_color='b8dc70', end_color='b8dc70', fill_type='solid')
        ws.cell(2*(n)+7,j+2).alignment = align


#___print Euro Put tree___(Blue)

        
ws.cell(2*n+9,1, 'Euro Put Payoff Tree')
ws.cell(2*n+9, 1).font = Font(bold=True)
for i in range(1, n+3):
    ws.cell(2*n+9, i).fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
ws.cell(3*n+11, 1).fill = PatternFill(start_color='d0ecfd', end_color='d0ecfd', fill_type='solid')
for j in range(n+1):
    for i in range(j+1):
        if j == 1:
            for k in range(n+1):
                ws.cell(3*(n)+9-k+1,j,k)
                ws.cell(3*(n)+9-k+1,j).fill = PatternFill(start_color='d0ecfd', end_color='d0ecfd', fill_type='solid')
                ws.cell(3*(n)+9-k+1,j)
                
        ws.cell(3*(n)+9-i+1,j+2,round(EPpayoff[i][j],4))
        ws.cell(3*(n)+9-i+1,j+2).alignment = align
        
        ws.cell(3*(n)+11,j+2,j)
        ws.cell(3*(n)+11,j+2).fill = PatternFill(start_color='d0ecfd', end_color='d0ecfd', fill_type='solid')
        ws.cell(3*(n)+11,j+2).alignment = align


#___print American Call tree___(Yellow)

        
ws.cell(3*n+13,1, 'American Call Payoff Tree')
ws.cell(3*n+13, 1).font = Font(bold=True)
for i in range(1, n+3):
    ws.cell(3*n+13, i).fill = PatternFill(start_color='ffd500', end_color='ffd500', fill_type='solid')
ws.cell(4*n+15, 1).fill = PatternFill(start_color='ffea80', end_color='ffea80', fill_type='solid')
for j in range(n+1):
    for i in range(j+1):
        if j == 1:
            for k in range(n+1):
                ws.cell(4*n+13-k+1,j,k)
                ws.cell(4*n+13-k+1,j).fill = PatternFill(start_color='ffea80', end_color='ffea80', fill_type='solid')
                ws.cell(4*n+13-k+1,j)
                
        ws.cell(4*n+13-i+1,j+2,round(ACpayoff[i][j],4))
        ws.cell(4*n+13-i+1,j+2).alignment = align
        
        ws.cell(4*(n)+15,j+2,j)
        ws.cell(4*(n)+15,j+2).fill = PatternFill(start_color='ffea80', end_color='ffea80', fill_type='solid')
        ws.cell(4*(n)+15,j+2).alignment = align


#___print American Put tree___(Purple)


ws.cell(4*n+17,1, 'American Put Payoff Tree')
ws.cell(4*n+17, 1).font = Font(bold=True)
for i in range(1, n+3):
    ws.cell(4*n+17, i).fill = PatternFill(start_color='bc9aff', end_color='bc9aff', fill_type='solid')
ws.cell(5*n+19, 1).fill = PatternFill(start_color='decdff', end_color='decdff', fill_type='solid')
for j in range(n+1):
    for i in range(j+1):
        if j == 1:
            for k in range(n+1):
                ws.cell(5*n+17-k+1,j,k)
                ws.cell(5*n+17-k+1,j).fill = PatternFill(start_color='decdff', end_color='decdff', fill_type='solid')
                ws.cell(5*n+17-k+1,j)
                
        ws.cell(5*n+17-i+1,j+2,round(APpayoff[i][j],4))
        ws.cell(5*n+17-i+1,j+2).alignment = align
        
        ws.cell(5*n+19,j+2,j)
        ws.cell(5*n+19,j+2).fill = PatternFill(start_color='decdff', end_color='decdff', fill_type='solid')
        ws.cell(5*n+19,j+2).alignment = align


#___save the xlxs file___             
wb.save('Binomial Option Pricing Model.xlsx')


##ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n+2)##ws.cell(1, 1).alignment = align
##ws.merge_cells(start_row=n+5, start_column=1, end_row=n+5, end_column=n+2)##ws.cell(n+5, 1).alignment = align
##ws.merge_cells(start_row=2*n+9, start_column=1, end_row=2*n+9, end_column=n+2)##ws.cell(2*n+9, 1).alignment = align
